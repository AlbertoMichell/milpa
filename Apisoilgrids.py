#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Apisoilgrids_xlsx.py
--------------------
Extrae variables edafológicas de SoilGrids (ISRIC v2.0) para puntos y profundidades,
y guarda el resultado en un archivo Excel (.xlsx).

Características:
- Lee CSV de entrada con separador , ; o tab (auto-detección).
- Encabezados requeridos: location, lat, lon, depth_cm_from, depth_cm_to
- Profundidades válidas SoilGrids (cm): 0-5, 5-15, 15-30, 30-60, 60-100, 100-200
- Pide estadísticas Q0.05, Q0.5 (mediana), Q0.95 y mean
- Parser robusto contra esquemas antiguos (M/mean) y actuales (depths -> values -> Q0.5)
- Genera .xlsx con hoja SoilGrids_Data (autofiltro, freeze panes, anchos de columna)
- Columnas extra de diagnóstico: _error (si falló la fila), _raw_json (si --debug)

Uso:
  python Apisoilgrids_xlsx.py --input Variablesedafologicas.csv --output dataset_edafologicas.xlsx [--sleep 0.5] [--debug]

Requisitos:
  pip install requests openpyxl
"""

import csv
import json
import time
import argparse
from pathlib import Path
from typing import Dict, Any, List

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

SOILGRIDS_POINT_API = "https://rest.isric.org/soilgrids/v2.0/properties/query"
SOILGRIDS_VARS = ["phh2o", "soc", "nitrogen", "sand", "silt", "clay", "bdod", "cec", "cfvo"]

# Columna de salida -> (variable SoilGrids, orden de preferencia de estadísticos)
OUT_COLS = {
    "ph_h2o": ("phh2o", ["Q0.5", "M", "mean"]),
    "soc_g_per_kg": ("soc", ["Q0.5", "M", "mean"]),
    "nitrogen_g_per_kg": ("nitrogen", ["Q0.5", "M", "mean"]),
    "sand_pct": ("sand", ["Q0.5", "M", "mean"]),
    "silt_pct": ("silt", ["Q0.5", "M", "mean"]),
    "clay_pct": ("clay", ["Q0.5", "M", "mean"]),
    "bulk_density_bdod_kg_per_m3": ("bdod", ["Q0.5", "M", "mean"]),
    "cec_cmolc_per_kg": ("cec", ["Q0.5", "M", "mean"]),
    "coarse_fragments_vol_pct": ("cfvo", ["Q0.5", "M", "mean"]),
}

VALID_LAYERS = [(0,5), (5,15), (15,30), (30,60), (60,100), (100,200)]

def depth_to_soilgrids_layer(z_from: float, z_to: float) -> str:
    layer = (int(z_from), int(z_to))
    if layer not in VALID_LAYERS:
        raise ValueError(f"Depth layer {layer} is invalid. Use one of: {VALID_LAYERS}")
    return f"{int(z_from)}-{int(z_to)}cm"

def sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[',',';','\t'])
        return dialect.delimiter
    except Exception:
        return ','

def normalize_headers(headers: List[str]) -> List[str]:
    norm = []
    for h in headers:
        if h is None:
            norm.append(h)
        else:
            norm.append(h.strip())
    return norm

def fetch_soilgrids_point(lat: float, lon: float, depth_key: str) -> Dict[str, Any]:
    params = {
        "lat": lat,
        "lon": lon,
        "property": ",".join(SOILGRIDS_VARS),
        "depth": depth_key,
        "value": "Q0.05,Q0.5,Q0.95,mean,M"
    }
    headers = {"User-Agent": "milpa/1.0 (+https://milpa.local)"}
    r = requests.get(SOILGRIDS_POINT_API, params=params, headers=headers, timeout=60)
    r.raise_for_status()
    return r.json()

def parse_soilgrids_json(js: Dict[str, Any], depth_key: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    props = js.get("properties", {}) or {}

    for out_col, (var, stat_list) in OUT_COLS.items():
        val = None
        node = props.get(var)
        if isinstance(node, dict):
            # Nuevo esquema
            depths = node.get("depths")
            if isinstance(depths, dict):
                d = depths.get(depth_key)
                if isinstance(d, dict):
                    values = d.get("values")
                    if isinstance(values, dict):
                        for stat in stat_list:
                            if stat in values and values[stat] is not None:
                                val = values[stat]
                                break
            # Esquema antiguo (flat)
            if val is None:
                for stat in stat_list:
                    arr = node.get(stat)
                    if isinstance(arr, list) and arr:
                        val = arr[0]
                        break
        out[out_col] = val
    return out

def set_column_widths(ws: Worksheet):
    # Auto-anchos aproximados
    for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        header = col[0].value or ""
        if len(header) <= 12:
            width = 14
        elif len(header) <= 25:
            width = 22
        else:
            width = 36
        ws.column_dimensions[get_column_letter(i)].width = width

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="CSV con columnas: location, lat, lon, depth_cm_from, depth_cm_to")
    ap.add_argument("--output", required=True, help="Ruta .xlsx de salida")
    ap.add_argument("--sleep", type=float, default=0.5, help="Segundos de espera entre llamadas a la API")
    ap.add_argument("--debug", action="store_true", help="Agrega columna _raw_json")
    args = ap.parse_args()

    inp = Path(args.input); outp = Path(args.output)

    # Leer sample para detectar separador
    sample = inp.read_text(encoding="utf-8", errors="ignore")[:2048]
    delim = sniff_delimiter(sample)

    with inp.open("r", encoding="utf-8", newline="") as f_in:
        reader = csv.reader(f_in, delimiter=delim)
        rows = list(reader)
        if not rows:
            raise SystemExit("[ERROR] CSV vacío.")
        header = normalize_headers(rows[0])
        body = rows[1:]
        idx = {h: i for i, h in enumerate(header)}

        required = ["location", "lat", "lon", "depth_cm_from", "depth_cm_to"]
        missing = [c for c in required if c not in idx]
        if missing:
            raise SystemExit(f"[ERROR] Faltan columnas requeridas: {missing}\nPresentes: {header}")

        # Preparar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "SoilGrids_Data"

        out_headers = header.copy()
        for col in list(OUT_COLS.keys()) + ["_error"]:
            if col not in out_headers:
                out_headers.append(col)
        if args.debug and "_raw_json" not in out_headers:
            out_headers.append("_raw_json")

        ws.append(out_headers)

        # Procesar filas
        for r in body:
            row = {h: (r[idx[h]] if idx[h] < len(r) else "") for h in header}
            # Limpieza
            if "_error" in row: del row["_error"]
            if "_raw_json" in row: del row["_raw_json"]

            try:
                lat = float(row["lat"]); lon = float(row["lon"])
                z_from = float(row["depth_cm_from"]); z_to = float(row["depth_cm_to"])
                depth_key = depth_to_soilgrids_layer(z_from, z_to)

                js = fetch_soilgrids_point(lat, lon, depth_key)
                values = parse_soilgrids_json(js, depth_key)

                if args.debug:
                    row["_raw_json"] = json.dumps(js)[:5000]

                for k in OUT_COLS.keys():
                    row[k] = values.get(k)
            except Exception as e:
                row["_error"] = str(e)

            ws.append([row.get(h, "") for h in out_headers])
            time.sleep(args.sleep)

        # Estética básica
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        set_column_widths(ws)

        # Guardar
        outp.parent.mkdir(parents=True, exist_ok=True)
        wb.save(outp)

if __name__ == "__main__":
    main()
