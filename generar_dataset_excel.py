"""
Genera un archivo Excel sintetico para el sistema MILPA.
Hojas: cultivos, sensores, global
Uso: py -3 generar_dataset_excel.py
Requiere: pip install openpyxl
"""
import os
import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).parent / "dataset_sintetico_milpa.xlsx"

# ──────────────────────────────────────────────────────────────
# Datos de cultivos
# ──────────────────────────────────────────────────────────────
CULTIVOS = [
    {
        "crop_name": "maiz",
        "variety": "Criollo",
        "planted_at": "2026-01-15",
        "expected_harvest_at": "2026-05-20",
        "growth_stage": "desarrollo",
        "status": "activo",
        "progress": 65,
        "notes": "Lote norte, riego por goteo",
    },
    {
        "crop_name": "frijol",
        "variety": "Negro",
        "planted_at": "2026-02-01",
        "expected_harvest_at": "2026-05-15",
        "growth_stage": "establecimiento",
        "status": "activo",
        "progress": 40,
        "notes": "Lote sur, sin riego artificial",
    },
    {
        "crop_name": "chile",
        "variety": "Serrano",
        "planted_at": "2026-02-10",
        "expected_harvest_at": "2026-06-10",
        "growth_stage": "siembra",
        "status": "activo",
        "progress": 20,
        "notes": "Invernadero, zona protegida",
    },
    {
        "crop_name": "calabaza",
        "variety": None,
        "planted_at": "2026-02-20",
        "expected_harvest_at": "2026-06-30",
        "growth_stage": "siembra",
        "status": "activo",
        "progress": 10,
        "notes": "Siembra directa en campo abierto",
    },
    {
        "crop_name": "tomate",
        "variety": "Saladette",
        "planted_at": "2026-01-25",
        "expected_harvest_at": "2026-05-10",
        "growth_stage": "maduracion",
        "status": "activo",
        "progress": 80,
        "notes": "Lote este, fertirrigacion semanal",
    },
]

# ──────────────────────────────────────────────────────────────
# Generar lecturas semanales de sensores (24 semanas)
# ──────────────────────────────────────────────────────────────
BASE = datetime.datetime(2025, 9, 1, 12, 0, 0)
WEEKS = 24
INTERVAL = 7  # dias

SENSOR_COLS = ["crop_name", "soil_moisture", "air_temp", "air_humidity", "light", "precipitation", "wind_speed", "created_at"]

sensor_rows = []
for week in range(WEEKS):
    ts = BASE + datetime.timedelta(days=week * INTERVAL)
    ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
    for ci, cultivo in enumerate(CULTIVOS):
        seed = week + ci * 3
        soil    = round(44 + (seed % 8) * 3.2, 1)
        air_t   = round(19 + (seed % 9) * 1.7, 1)
        air_h   = round(42 + (seed % 10) * 3.1, 1)
        light   = round(58 + (seed % 9) * 3.2, 1)
        precip  = round(8.5 if seed % 5 == 0 else 0.0, 1)
        wind    = round(5.5 + (seed % 6) * 0.85, 1)
        sensor_rows.append({
            "crop_name":     cultivo["crop_name"],
            "soil_moisture": soil,
            "air_temp":      air_t,
            "air_humidity":  air_h,
            "light":         light,
            "precipitation": precip,
            "wind_speed":    wind,
            "created_at":    ts_str,
        })

# ──────────────────────────────────────────────────────────────
# Lecturas globales edafológicas (semanales, misma escala)
# ──────────────────────────────────────────────────────────────
GLOBAL_COLS = ["location_name","soil_temp","air_temp","air_humidity","soil_moisture","precipitation","wind_speed","ph","conductivity","notes","created_at"]
LOCATIONS = ["Region Norte", "Region Sur", "Region Centro"]

global_rows = []
for week in range(WEEKS):
    ts = BASE + datetime.timedelta(days=week * INTERVAL)
    ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
    for li, loc in enumerate(LOCATIONS):
        seed = week + li * 7
        global_rows.append({
            "location_name": loc,
            "soil_temp":     round(17 + (seed % 7) * 1.1, 1),
            "air_temp":      round(20 + (seed % 10) * 1.4, 1),
            "air_humidity":  round(45 + (seed % 9) * 2.8, 1),
            "soil_moisture": round(40 + (seed % 10) * 2.9, 1),
            "precipitation": round(9.0 if seed % 4 == 0 else 0.0, 1),
            "wind_speed":    round(6 + (seed % 6) * 0.7, 1),
            "ph":            round(6.2 + (seed % 5) * 0.12, 2),
            "conductivity":  round(0.9 + (seed % 4) * 0.15, 2),
            "notes":         "Lectura edafologica semanal automatica",
            "created_at":    ts_str,
        })

# ──────────────────────────────────────────────────────────────
# Estilos
# ──────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="2E7D32")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL    = PatternFill("solid", fgColor="F1F8E9")
BORDER_SIDE   = Side(style="thin", color="BDBDBD")
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT    = Alignment(horizontal="left", vertical="center")


def style_sheet(ws, col_names, rows_data):
    """Aplica encabezados con estilo y rellena datos."""
    # Encabezados
    for ci, col in enumerate(col_names, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = ALIGN_CENTER
    # Datos
    for ri, row in enumerate(rows_data, start=2):
        fill = ZEBRA_FILL if ri % 2 == 0 else None
        for ci, col in enumerate(col_names, start=1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col))
            cell.border = CELL_BORDER
            if fill:
                cell.fill = fill
            cell.alignment = ALIGN_LEFT
    # Ancho automático
    for ci, col in enumerate(col_names, start=1):
        max_len = len(col)
        for row in rows_data:
            val = str(row.get(col, ""))
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)
    # Fijar fila de encabezado
    ws.freeze_panes = ws["A2"]


# ──────────────────────────────────────────────────────────────
# Construir libro
# ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # quitar hoja por defecto

# Hoja cultivos
ws_c = wb.create_sheet("cultivos")
CULT_COLS = ["crop_name","variety","planted_at","expected_harvest_at","growth_stage","status","progress","notes"]
style_sheet(ws_c, CULT_COLS, CULTIVOS)

# Hoja sensores
ws_s = wb.create_sheet("sensores")
style_sheet(ws_s, SENSOR_COLS, sensor_rows)

# Hoja global
ws_g = wb.create_sheet("global")
style_sheet(ws_g, GLOBAL_COLS, global_rows)

# Hoja README
ws_r = wb.create_sheet("README")
ws_r["A1"] = "Dataset Sintético MILPA"
ws_r["A1"].font = Font(bold=True, size=14, color="2E7D32")
ws_r["A3"] = f"Generado el: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
ws_r["A4"] = f"Cultivos: {len(CULTIVOS)}"
ws_r["A5"] = f"Lecturas de sensores: {len(sensor_rows)} ({WEEKS} semanas x {len(CULTIVOS)} cultivos)"
ws_r["A6"] = f"Lecturas globales: {len(global_rows)} ({WEEKS} semanas x {len(LOCATIONS)} ubicaciones)"
ws_r["A8"] = "Instrucciones:"
ws_r["A8"].font = Font(bold=True)
ws_r["A9"]  = "1. Abre http://localhost:8080/ui/ingesta"
ws_r["A10"] = "2. Arrastra este archivo al panel de carga"
ws_r["A11"] = "3. Selecciona el usuario destino"
ws_r["A12"] = "4. Haz clic en 'Importar'"
ws_r["A14"] = "Hojas requeridas: cultivos | sensores | global"
ws_r.column_dimensions["A"].width = 60

wb.save(OUTPUT)
print(f"✅ Dataset generado: {OUTPUT}")
print(f"   Cultivos : {len(CULTIVOS)}")
print(f"   Sensores : {len(sensor_rows)} filas")
print(f"   Global   : {len(global_rows)} filas")
print(f"\n→ Cargalo en: http://localhost:8080/ui/ingesta")
